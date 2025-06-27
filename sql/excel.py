
import csv, os
import openpyxl
from pathlib import Path
root_path=Path('C:/Users/Zeina/Downloads/Automate/automate_online-materials/excelSpreadsheets')
excelFiles=os.listdir(root_path)
for wb in excelFiles:
    if not wb.endswith('.xlsx'):
        continue
    workbook = openpyxl.load_workbook(root_path / wb)
    workbook_name =str(wb)
    workbook_name_no_ext=os.path.splitext(workbook_name)[0]# التخلص من امتداد الملف

    excelFiles_name= workbook_name_no_ext+"_"
    sheet_names = workbook.sheetnames
    for sheet in sheet_names:
        my_sheet =workbook[sheet]
        CSVfile = open(root_path/'CSV'/Path(excelFiles_name+sheet+".csv"),'w', newline="")
        writer = csv.writer(CSVfile)
        data = []
        for i in range(1, my_sheet.max_row+ 1):
            for j in range(1, my_sheet.max_column + 1):
                data.append(my_sheet.cell(row=i, column=j).value)
            writer.writerow(data)
        CSVfile.close()
